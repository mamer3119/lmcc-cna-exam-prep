"""Build LMCC CNA skills workbook — one sheet per skill, OPEN/CORE/CLOSE sections."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
PAYLOAD = ROOT / "exports" / "skills-xlsm-payload.json"
DEFAULT_OUT = ROOT / "exports" / "LMCC-CNA-Skills-Tagged.xlsx"

HEADERS = [
    "Step #",
    "Step Text",
    "Boilerplate Tag ID",
    "Detailed Tag Text",
    "Tag Category",
    "Phase Word",
    "Clinical Note",
    "Sub-Steps",
    "Critical Category",
    "Exam Scorecard",
]

SECTION_FILLS = {
    "OPEN": PatternFill("solid", fgColor="E8F4EA"),
    "CORE": PatternFill("solid", fgColor="EAF0FA"),
    "CLOSE": PatternFill("solid", fgColor="FCEFE8"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14)
META_FONT = Font(name="Arial", size=10, color="374151")
BODY_FONT = Font(name="Arial", size=10)
SECTION_FONT = Font(name="Arial", bold=True, size=11)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def safe_sheet_name(study_order: int, title: str) -> str:
    base = f"S{study_order:02d} {title}"
    base = re.sub(r"[\[\]\*\?:/\\]", " ", base)
    return base[:31]


def set_col_widths(ws) -> None:
    widths = [8, 62, 22, 62, 16, 14, 28, 40, 18, 32]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_index(wb: Workbook, payload: dict) -> None:
    ws = wb.active
    ws.title = "Index"
    ws["A1"] = "LMCC CNA Skills — Tagged Checklists"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {payload['generatedAt']}"
    ws["A2"].font = META_FONT
    ws["A3"] = payload.get("pathwayTagline", "")
    ws["A3"].font = META_FONT

    index_headers = [
        "Study #",
        "Exam Skill #",
        "Sheet Name",
        "Title",
        "Section",
        "Template",
        "Steps",
    ]
    row = 5
    for col, label in enumerate(index_headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER

    for skill in payload["skills"]:
        row += 1
        sheet_name = safe_sheet_name(skill["studyOrder"], skill["title"])
        values = [
            skill["studyOrder"],
            skill["examSkillNumber"],
            sheet_name,
            skill["title"],
            skill["section"],
            skill["template"],
            skill["stepCount"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 8


def write_skill_sheet(wb: Workbook, skill: dict, segment_labels: dict) -> None:
    ws = wb.create_sheet(title=safe_sheet_name(skill["studyOrder"], skill["title"]))

    ws.merge_cells("A1:J1")
    ws["A1"] = skill["title"]
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")

    meta = (
        f"{skill['examCardLabel']}  |  Section: {skill['section']}  |  "
        f"Template: {skill['template']}  |  Module: {skill['moduleVerb']}  |  "
        f"{skill['stepCount']} steps"
    )
    ws.merge_cells("A2:J2")
    ws["A2"] = meta
    ws["A2"].font = META_FONT

    row = 4
    for col, label in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    segment_order = ["open", "core", "close"]
    steps_by_segment = {seg: [] for seg in segment_order}
    for step in skill["steps"]:
        seg = step.get("segment", "core")
        steps_by_segment.setdefault(seg, []).append(step)

    for seg in segment_order:
        section_steps = steps_by_segment.get(seg, [])
        if not section_steps:
            continue

        row += 1
        label = segment_labels.get(seg, seg.upper())
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=len(HEADERS)
        )
        section_cell = ws.cell(row=row, column=1, value=f"── {label} ──")
        section_cell.font = SECTION_FONT
        section_cell.fill = SECTION_FILLS.get(label, PatternFill())
        section_cell.alignment = Alignment(horizontal="left", vertical="center")

        for step in section_steps:
            row += 1
            sub_steps = " | ".join(step.get("subSteps") or [])
            values = [
                step["id"],
                step["text"],
                step.get("boilerplateId") or "",
                step.get("detailedTagText") or "",
                step.get("tagCategoryLabel") or "",
                step.get("phaseWord") or "",
                step.get("note") or "",
                sub_steps,
                step.get("criticalCategory") or "",
                step.get("examScorecard") or "",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A5"
    set_col_widths(ws)


def main() -> int:
    payload_path = Path(sys.argv[1]) if len(sys.argv) > 1 else PAYLOAD
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    if not payload_path.is_file():
        print(f"Missing payload: {payload_path}", file=sys.stderr)
        print(
            "Run: node --experimental-strip-types scripts/export-skills-xlsm-data.mjs",
            file=sys.stderr,
        )
        return 1

    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    segment_labels = payload.get(
        "segmentLabels", {"open": "OPEN", "core": "CORE", "close": "CLOSE"}
    )

    wb = Workbook()
    write_index(wb, payload)
    for skill in payload["skills"]:
        write_skill_sheet(wb, skill, segment_labels)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path} ({len(payload['skills'])} skill sheets + Index)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
